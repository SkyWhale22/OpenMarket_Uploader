from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


from Uploader.UploaderBase import UploaderBase
from openpyxl import load_workbook
from Items.Base import ItemBase

from Items.ESM import ItemESM
from Items.ESM import ESMItemInfo
from Items.ESM import Languages



"""
	ESMUploader class definition
"""
class ESMUploader:
	####################################################################################################
	#                                        Member Variables                                          #
	####################################################################################################

	# -------------------------------------------------------------------------------------------------
	# Public
	# -------------------------------------------------------------------------------------------------
	m_itemList = None
	m_driver = None

	####################################################################################################
	#                                        Member Functions                                          #
	####################################################################################################

	# -------------------------------------------------------------------------------------------------
	# Private
	# -------------------------------------------------------------------------------------------------
	def __init__(self):
		print("------- Init! ------")
		self.m_itemList = []
		self.m_path = "C:/Users/kimsu/OneDrive/Documents/Development/chromedriver_win32/chromedriver.exe"
		self.m_driver = webdriver.Chrome(self.m_path)
		self.m_driver.get('https://www.esmplus.com/Member/SignIn/LogOn')

	# -------------------------------------------------------------------------------------------------
	# Public
	# -------------------------------------------------------------------------------------------------

	def UpLoad(self):
		# Step 1: Sign in to the ESM Plus.
		self.SignIn("hymedi682", "rlawnsgh12**")

		# Step 2: Get data from .xlsx file and contain these date in the the list.
		self.GetDataFromExel()

		# Step 3: To make it sure, check read data.
		self.PrintAllItems()

		# Step 4: While m_itemList contains items, loop the following steps.
		while(len(self.m_itemList) > 0):
			# Step 5: Get the last item from the list, and remove it from it.
			#         You are able to get an item from head by changing code below 'item = self.m_itemList.pop(0)
			item = self.m_itemList.pop()

			# Step 6: Make the chrome browser getting into item registeration page.
			self.GetInItemRegisterView()

			# Step 7: Fill out basic information of the item which will be uploaded.
			self.FillOutBasicInfo(item)
			# Step 8: Fill out exposure information of the item which will be uploaded.
			self.FillOutExposureInfo(item)
			# Step 9: Fill out additional information of the item which will be uploaded.
			self.FillOutAdditionalInfo(item)

			# Step 10: The last step. Click uploading buttton.
			self.ClickSubmitBtn()

	# SignIn logic
	def SignIn(self, id, pw):
		signInID    = self.m_driver.find_element_by_id("Id")
		signInPW    = self.m_driver.find_element_by_id("Password")
		signInBtn   = self.m_driver.find_element_by_id("btnLogOn")

		signInID.clear()
		signInPW.clear()

		signInID.send_keys(id)
		signInPW.send_keys(pw)

		signInBtn.click()

	# Print All item's data on the console.
	# Make sure to comment following code out so that it won't harm the performance.
	# Or make the visual console.
	def PrintAllItems(self):
		print("이름\t \t \t|\t가격\t \t|\t제품코드	\t|\t재고수량\t|\t브랜드\t|\t프로모션이름\t|\t원산지(국내, 수입)\t|\t원산지(국내산->도, 수입산->대륙)\t|\t원산지(국내산->지역 이름(시, 군, 읍), 수입산->국가)\t|\t\t 영문 이름 \t\t|\t\t 일문이름 \t\t|\t\t 중문이름 \t\t|\t \t \t 취급시 주의사항 \t \t \t|\t\t\t\t 품질 보증기준\t\t\t|\t A/S책임자 \t|\t주문후 예상 배송기간\t|\t의료기기 법상 허가번호\t|\t KC인증 필 유무\t|\t정격전압/소비전력\t|\t동일모델의 출시년월	\t|\t제조자/수입자\t|\t제조국\t|\t	제품의 사용목적 및 사용방법")

		for i in range(0, len(self.m_itemList)):
			item = self.m_itemList[i]
			#item = self.m_itemList["testName"]

			# print(item.GetProductName() + "\t  \t" + str(item.GetProductPrice()) + "\t  \t" + item.GetProductCode() + "\t  \t" + str(item.GetNumberOfStocks()) + "\t  \t \t" +
			#       item.GetProductBrandName() + "\t  \t" + item.GetPromotionName() + "\t  \t" + item.GetSourceOfOrigin() + "\t  \t \t \t \t \t \t \t" + item.GetSourceOfOrigin_Sub1() +
			#       "\t \t \t \t \t \t \t \t \t \t" + item.GetSourceOfOrigin_Sub2() + "\t\t\t\t\t\t\t\t\t\t" + item.GetMultiLingProductName(Languages.KEnglish) + "\t\t\t\t" +
			#       item.GetMultiLingProductName(Languages.KJapanese) + "\t\t\t\t" + item.GetMultiLingProductName(Languages.KChinese) + "\t\t\t\t" + item.GetESMItemInfo().m_caution +
			#       "\t\t\t" + item.GetESMItemInfo().m_warrenty + "\t\t\t" + item.GetESMItemInfo().m_asRepresentitive + "\t\t\t\t" + item.GetESMItemInfo().m_estimatedDeliveryDays +
			#       "\t\t\t\t" + str(item.GetESMItemInfo().m_licenceNumber) + "\t\t\t\t\t\t" + str(item.GetESMItemInfo().m_kcCredit) + "\t\t\t\t" + item.GetESMItemInfo().m_ratedVoltage  +
			#       "\t\t\t\t\t" + str(item.GetESMItemInfo().m_releaseDate) + "\t\t\t\t" + item.GetESMItemInfo().m_manufacturer  + "\t\t" + item.GetESMItemInfo().m_originCountry  + "\t\t\t\t" +
			#       item.GetESMItemInfo().m_howToUse)

	# Get in item registration view
	def GetInItemRegisterView(self):
		self.m_driver.find_element_by_xpath('//*[@id="TDM001"]').click()
		self.m_driver.find_element_by_xpath('//*[@id="TDM395"]').click()

		self.m_driver.implicitly_wait(1.5)
		self.m_driver.switch_to.frame(self.m_driver.find_element_by_id("ifm_TDM395"))

	# Fill out basic information of item which will be uploaded
	def FillOutBasicInfo(self, item):
		# Wait until the page completely rendered
		self.m_driver.implicitly_wait(1.5)

		# Always check is the field empty.
		if(self.IsAttributeNone(item.GetProductName())):
			self.m_driver.find_element_by_id("txtGoodsNameSearch").send_keys(item.GetProductName())

		if(self.IsAttributeNone(item.GetPromotionName())):
			self.m_driver.find_element_by_id("txtGoodsNamePrmt").send_keys(item.GetPromotionName())

		if(self.IsAttributeNone(item.GetMultiLingProductName(Languages.KEnglish))):
			self.m_driver.find_element_by_id("txtGoodsNameEng2").send_keys(item.GetMultiLingProductName(Languages.KEnglish))

		if(self.IsAttributeNone(item.GetMultiLingProductName(Languages.KChinese))):
			self.m_driver.find_element_by_id("txtGoodsNameChn2").send_keys(item.GetMultiLingProductName(Languages.KChinese))

		if(self.IsAttributeNone(item.GetMultiLingProductName(Languages.KJapanese))):
			self.m_driver.find_element_by_id("txtGoodsNameJpn2").send_keys(item.GetMultiLingProductName(Languages.KJapanese))

		if(self.IsAttributeNone(item.GetProductPrice())):
			self.m_driver.find_element_by_id("txtGoodsPrice").send_keys(item.GetProductPrice())

		if(self.IsAttributeNone( item.GetSourceOfOrigin().GetFirstXpath())):
			self.m_driver.find_element_by_xpath(item.GetSourceOfOrigin().GetFirstXpath()).click()
			if (self.IsAttributeNone(item.GetSourceOfOrigin().GetSecondXpath())):
				self.m_driver.implicitly_wait(10)
				fix1 = item.GetSourceOfOrigin().GetSecondXpath()
				fix1 = fix1[1:len(item.GetSourceOfOrigin().GetSecondXpath())]
				#self.m_driver.find_element_by_xpath('//*[@id="sltOriginLName"]/option[5]').click()

				self.m_driver.find_element_by_id("sltOriginLName.selectType01.native").__setattr__('value', '강원')

				print(item.GetSourceOfOrigin().GetFirstXpath() + item.GetSourceOfOrigin().GetSecondXpath())
				# if (self.IsAttributeNone(item.GetSourceOfOrigin().GetThirdXpath())):
				# 	self.m_driver.implicitly_wait(2)
				# 	self.m_driver.find_element_by_xpath(item.GetSourceOfOrigin().GetThirdXpath()).click()
				# 	print(item.GetSourceOfOrigin().GetFirstXpath() + item.GetSourceOfOrigin().GetSecondXpath() + item.GetSourceOfOrigin().GetThirdXpath())

	# sltOriginLName > option:nth-child(3)
	# Fill out exposure information of item which will be uploaded
	def FillOutExposureInfo(self, item):
		self.m_driver.find_element_by_class_name("menu_item2").click()

		self.m_driver.find_element_by_xpath('//*[@id="ddlOfficialNotiTemplateList"]/option[2]').click()

		if(self.IsAttributeNone(item.GetProductName())):
			self.m_driver.find_element_by_xpath('//*[@id="16-1"]/textarea').clear()
			self.m_driver.find_element_by_xpath('//*[@id="16-1"]/textarea').send_keys(item.GetProductName())

		if(self.IsAttributeNone(item.GetESMItemInfo().m_caution)):
			self.m_driver.find_element_by_xpath('//*[@id="16-10"]/textarea').clear()
			self.m_driver.find_element_by_xpath('//*[@id="16-10"]/textarea').send_keys(item.GetESMItemInfo().m_caution)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_licenceNumber)):
			self.m_driver.find_element_by_xpath('//*[@id="16-2"]/textarea').send_keys(item.GetESMItemInfo().m_licenceNumber)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_kcCredit)):
			self.m_driver.find_element_by_xpath('//*[@id="16-4"]/textarea').send_keys(item.GetESMItemInfo().m_kcCredit)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_ratedVoltage)):
			self.m_driver.find_element_by_xpath('//*[@id="16-5"]/textarea').send_keys(item.GetESMItemInfo().m_ratedVoltage)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_releaseDate)):
			self.m_driver.find_element_by_xpath('//*[@id="16-6"]/textarea').clear()
			self.m_driver.find_element_by_xpath('//*[@id="16-6"]/textarea').send_keys(item.GetESMItemInfo().m_releaseDate)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_manufacturer)):
			self.m_driver.find_element_by_xpath('//*[@id="16-7"]/textarea').clear()
			self.m_driver.find_element_by_xpath('//*[@id="16-7"]/textarea').send_keys(item.GetESMItemInfo().m_manufacturer)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_originCountry)):
			self.m_driver.find_element_by_xpath('//*[@id="16-8"]/textarea').clear()
			self.m_driver.find_element_by_xpath('//*[@id="16-8"]/textarea').send_keys(item.GetESMItemInfo().m_originCountry)

		if(self.IsAttributeNone(item.GetESMItemInfo().m_howToUse)):
			self.m_driver.find_element_by_xpath('//*[@id="16-9"]/textarea').clear()
			self.m_driver.find_element_by_xpath('//*[@id="16-9"]/textarea').send_keys(item.GetESMItemInfo().m_howToUse)

		self.m_driver.find_element_by_id("rdoCommonDeliveryWayOPTSEL1").click()
		self.m_driver.find_element_by_xpath('//*[@id="selCommonDeliveryWayOPTSELParcelCOMP"]/option[7]').click()

	# Fill out additional information of the ite m which will be uploaded.
	def FillOutAdditionalInfo(self, item):
		self.m_driver.find_element_by_class_name("menu_item3").click()

		if(self.IsAttributeNone(item.GetProductCode())):
			self.m_driver.find_element_by_id("txtItemCode").send_keys(item.GetProductCode())

	# Finally, click the submit button.
	def ClickSubmitBtn(self):
		self.m_driver.find_element_by_class_name("menu_item4").click()
		#self.m_driver.find_element_by_xpath('//*[@id="contents"]/div[4]/a[2]').click()

	def GetDataFromExel(self):
		itemList    = load_workbook(filename='테스트용.xlsx')
		sheet       = itemList['Sheet1']

		#for currentRow in sheet.iter_rows('A2:A' + str(sheet.max_row)):
		for i in range(2, sheet.max_row + 1, 1):
			name                    = sheet.cell(row=i, column=1).value
			price                   = sheet.cell(row=i, column=2).value
			modelCode               = sheet.cell(row=i, column=3).value
			stockNum                = sheet.cell(row=i, column=4).value
			brandName               = sheet.cell(row=i, column=5).value
			promotionName           = sheet.cell(row=i, column=6).value
			sourceOfOrigin          = sheet.cell(row=i, column=7).value
			sourceOfOrigin_sub1     = sheet.cell(row=i, column=8).value
			sourceOfOrigin_sub2     = sheet.cell(row=i, column=9).value
			engName                 = sheet.cell(row=i, column=10).value
			japName                 = sheet.cell(row=i, column=11).value
			chiName                 = sheet.cell(row=i, column=12).value
			caution                 = sheet.cell(row=i, column=13).value
			warrenty                = sheet.cell(row=i, column=14).value
			asRepresentitive        = sheet.cell(row=i, column=15).value
			estimatedDeliveryDays   = sheet.cell(row=i, column=16).value
			licenceNumber           = sheet.cell(row=i, column=17).value
			kcCredit                = sheet.cell(row=i, column=18).value
			ratedVoltage            = sheet.cell(row=i, column=19).value
			releaseDate             = sheet.cell(row=i, column=20).value
			manufacturer            = sheet.cell(row=i, column=21).value
			manufacturingCounty     = sheet.cell(row=i, column=7).value
			howToUse                = sheet.cell(row=i, column=23).value

			esmItemInfo = ESMItemInfo(
				name,
				caution,
				warrenty,
				asRepresentitive,
				estimatedDeliveryDays,
				licenceNumber,
				kcCredit,
				ratedVoltage,
				releaseDate,
				manufacturer,
				manufacturingCounty,
				howToUse
			)

			#item = ItemESM(name, price, modelCode, stockNum, brandName, promotionName, sourceOfOrigin, sourceOfOrigin_sub1, sourceOfOrigin_sub2,engName,japName,chiName,esmItemInfo)
			self.m_itemList.append(ItemESM(name, price, modelCode, stockNum, brandName, promotionName, sourceOfOrigin, sourceOfOrigin_sub1, sourceOfOrigin_sub2,engName,japName,chiName,esmItemInfo))

	# Value checker. This check is the variable empty or not.
	def IsAttributeNone(self, attribute):
		if(attribute == None):
			return False

		return True

	def InsertSourceOfOriginData(self, item):
		if (self.IsAttributeNone(item.GetSourceOfOrigin())):
			self.m_driver.find_element_by_xpath(item.GetSourceOfOrigin()).click()
			#item.



# For testing code.
# Make sure to comment this out or remove before releasing.
test = ESMUploader()
test.UpLoad()